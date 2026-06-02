"""
Report service - business logic for generating financial reports
"""
from sqlalchemy.orm import Session
from sqlalchemy import func, extract
from datetime import date, datetime, timedelta
from typing import Dict, List, Any
from collections import defaultdict

from app.models.expense import Expense
from app.models.income import Income

class ReportService:
    """Service class for report generation and analytics"""
    
    @staticmethod
    def generate_report(
        user_id: int,
        start_date: date,
        end_date: date,
        db: Session
    ) -> Dict[str, Any]:
        """
        Generate comprehensive financial report for date range
        
        Returns:
        - Summary statistics (total income, expenses, balance)
        - Category breakdown
        - Income source breakdown
        - Top expenses
        - Monthly trends
        - Daily averages
        """
        
        # Get all expenses in date range
        expenses = db.query(Expense).filter(
            Expense.user_id == user_id,
            Expense.date >= start_date,
            Expense.date <= end_date
        ).all()
        
        # Get all income in date range
        incomes = db.query(Income).filter(
            Income.user_id == user_id,
            Income.date >= start_date,
            Income.date <= end_date
        ).all()
        
        # Calculate summary statistics
        total_expenses = sum(e.amount for e in expenses)
        total_income = sum(i.amount for i in incomes)
        balance = total_income - total_expenses
        
        # Category breakdown
        category_breakdown = defaultdict(lambda: {"total": 0.0, "count": 0, "percentage": 0.0})
        for expense in expenses:
            category_breakdown[expense.category]["total"] += expense.amount
            category_breakdown[expense.category]["count"] += 1
        
        # Calculate percentages
        for category in category_breakdown:
            if total_expenses > 0:
                category_breakdown[category]["percentage"] = (
                    category_breakdown[category]["total"] / total_expenses
                ) * 100
        
        # Sort by total descending
        category_breakdown = dict(sorted(
            category_breakdown.items(),
            key=lambda x: x[1]["total"],
            reverse=True
        ))
        
        # Income source breakdown
        source_breakdown = defaultdict(lambda: {"total": 0.0, "count": 0, "percentage": 0.0})
        for income in incomes:
            source_breakdown[income.source]["total"] += income.amount
            source_breakdown[income.source]["count"] += 1
        
        # Calculate percentages
        for source in source_breakdown:
            if total_income > 0:
                source_breakdown[source]["percentage"] = (
                    source_breakdown[source]["total"] / total_income
                ) * 100
        
        # Sort by total descending
        source_breakdown = dict(sorted(
            source_breakdown.items(),
            key=lambda x: x[1]["total"],
            reverse=True
        ))
        
        # Top 10 expenses
        top_expenses = sorted(expenses, key=lambda x: x.amount, reverse=True)[:10]
        top_expenses_data = [
            {
                "id": e.id,
                "title": e.title,
                "amount": e.amount,
                "category": e.category,
                "date": e.date.isoformat(),
                "description": e.description
            }
            for e in top_expenses
        ]
        
        # Monthly trends
        monthly_trends = ReportService._calculate_monthly_trends(
            expenses, incomes, start_date, end_date
        )
        
        # Calculate daily averages
        days_in_period = (end_date - start_date).days + 1
        avg_daily_expense = total_expenses / days_in_period if days_in_period > 0 else 0
        avg_daily_income = total_income / days_in_period if days_in_period > 0 else 0
        
        # Identify highest spending category
        highest_category = max(
            category_breakdown.items(),
            key=lambda x: x[1]["total"]
        )[0] if category_breakdown else None
        
        # Identify highest income source
        highest_source = max(
            source_breakdown.items(),
            key=lambda x: x[1]["total"]
        )[0] if source_breakdown else None
        
        return {
            "period": {
                "start_date": start_date.isoformat(),
                "end_date": end_date.isoformat(),
                "days": days_in_period
            },
            "summary": {
                "total_income": round(total_income, 2),
                "total_expenses": round(total_expenses, 2),
                "balance": round(balance, 2),
                "income_count": len(incomes),
                "expense_count": len(expenses),
                "avg_daily_expense": round(avg_daily_expense, 2),
                "avg_daily_income": round(avg_daily_income, 2),
                "highest_category": highest_category,
                "highest_source": highest_source
            },
            "category_breakdown": {
                category: {
                    "total": round(data["total"], 2),
                    "count": data["count"],
                    "percentage": round(data["percentage"], 2)
                }
                for category, data in category_breakdown.items()
            },
            "source_breakdown": {
                source: {
                    "total": round(data["total"], 2),
                    "count": data["count"],
                    "percentage": round(data["percentage"], 2)
                }
                for source, data in source_breakdown.items()
            },
            "top_expenses": top_expenses_data,
            "monthly_trends": monthly_trends,
            "generated_at": datetime.utcnow().isoformat()
        }
    
    @staticmethod
    def _calculate_monthly_trends(
        expenses: List[Expense],
        incomes: List[Income],
        start_date: date,
        end_date: date
    ) -> List[Dict[str, Any]]:
        """Calculate monthly income and expense trends"""
        
        # Group by month
        monthly_data = defaultdict(lambda: {"income": 0.0, "expenses": 0.0})
        
        for expense in expenses:
            month_key = expense.date.strftime("%Y-%m")
            monthly_data[month_key]["expenses"] += expense.amount
        
        for income in incomes:
            month_key = income.date.strftime("%Y-%m")
            monthly_data[month_key]["income"] += income.amount
        
        # Convert to list and sort by month
        trends = []
        for month_key in sorted(monthly_data.keys()):
            data = monthly_data[month_key]
            trends.append({
                "month": month_key,
                "income": round(data["income"], 2),
                "expenses": round(data["expenses"], 2),
                "balance": round(data["income"] - data["expenses"], 2)
            })
        
        return trends
    
    @staticmethod
    def export_to_csv(report_data: Dict[str, Any]) -> str:
        """
        Convert report data to CSV format
        
        Returns CSV string with multiple sections:
        - Summary
        - Category breakdown
        - Income source breakdown
        - Top expenses
        - Monthly trends
        """
        import csv
        from io import StringIO
        
        output = StringIO()
        writer = csv.writer(output)
        
        # Header
        writer.writerow(["Financial Report"])
        writer.writerow(["Generated:", report_data["generated_at"]])
        writer.writerow(["Period:", f"{report_data['period']['start_date']} to {report_data['period']['end_date']}"])
        writer.writerow([])
        
        # Summary section
        writer.writerow(["SUMMARY"])
        writer.writerow(["Metric", "Value"])
        summary = report_data["summary"]
        writer.writerow(["Total Income", f"${summary['total_income']:.2f}"])
        writer.writerow(["Total Expenses", f"${summary['total_expenses']:.2f}"])
        writer.writerow(["Balance", f"${summary['balance']:.2f}"])
        writer.writerow(["Income Transactions", summary['income_count']])
        writer.writerow(["Expense Transactions", summary['expense_count']])
        writer.writerow(["Avg Daily Expense", f"${summary['avg_daily_expense']:.2f}"])
        writer.writerow(["Avg Daily Income", f"${summary['avg_daily_income']:.2f}"])
        writer.writerow([])
        
        # Category breakdown
        writer.writerow(["EXPENSE BY CATEGORY"])
        writer.writerow(["Category", "Total", "Count", "Percentage"])
        for category, data in report_data["category_breakdown"].items():
            writer.writerow([
                category,
                f"${data['total']:.2f}",
                data['count'],
                f"{data['percentage']:.2f}%"
            ])
        writer.writerow([])
        
        # Income source breakdown
        writer.writerow(["INCOME BY SOURCE"])
        writer.writerow(["Source", "Total", "Count", "Percentage"])
        for source, data in report_data["source_breakdown"].items():
            writer.writerow([
                source,
                f"${data['total']:.2f}",
                data['count'],
                f"{data['percentage']:.2f}%"
            ])
        writer.writerow([])
        
        # Top expenses
        writer.writerow(["TOP EXPENSES"])
        writer.writerow(["Title", "Amount", "Category", "Date", "Description"])
        for expense in report_data["top_expenses"]:
            writer.writerow([
                expense['title'],
                f"${expense['amount']:.2f}",
                expense['category'],
                expense['date'],
                expense.get('description', '')
            ])
        writer.writerow([])
        
        # Monthly trends
        writer.writerow(["MONTHLY TRENDS"])
        writer.writerow(["Month", "Income", "Expenses", "Balance"])
        for trend in report_data["monthly_trends"]:
            writer.writerow([
                trend['month'],
                f"${trend['income']:.2f}",
                f"${trend['expenses']:.2f}",
                f"${trend['balance']:.2f}"
            ])
        
        return output.getvalue()
    
    @staticmethod
    def export_to_pdf(report_data: Dict[str, Any]) -> bytes:
        """
        Convert report data to PDF format with charts and tables
        
        Returns PDF bytes
        """
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib import colors
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.enums import TA_CENTER, TA_RIGHT
        from io import BytesIO
        
        buffer = BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        elements = []
        styles = getSampleStyleSheet()
        
        # Custom styles
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            textColor=colors.HexColor('#8B5CF6'),
            spaceAfter=30,
            alignment=TA_CENTER
        )
        
        heading_style = ParagraphStyle(
            'CustomHeading',
            parent=styles['Heading2'],
            fontSize=16,
            textColor=colors.HexColor('#4B5563'),
            spaceAfter=12,
            spaceBefore=20
        )
        
        # Title
        elements.append(Paragraph("Financial Report", title_style))
        elements.append(Spacer(1, 0.2*inch))
        
        # Period info
        period_text = f"Period: {report_data['period']['start_date']} to {report_data['period']['end_date']} ({report_data['period']['days']} days)"
        elements.append(Paragraph(period_text, styles['Normal']))
        elements.append(Paragraph(f"Generated: {report_data['generated_at']}", styles['Normal']))
        elements.append(Spacer(1, 0.3*inch))
        
        # Summary section
        elements.append(Paragraph("Summary", heading_style))
        summary = report_data['summary']
        summary_data = [
            ['Metric', 'Value'],
            ['Total Income', f"${summary['total_income']:,.2f}"],
            ['Total Expenses', f"${summary['total_expenses']:,.2f}"],
            ['Balance', f"${summary['balance']:,.2f}"],
            ['Income Transactions', str(summary['income_count'])],
            ['Expense Transactions', str(summary['expense_count'])],
            ['Avg Daily Expense', f"${summary['avg_daily_expense']:,.2f}"],
            ['Avg Daily Income', f"${summary['avg_daily_income']:,.2f}"],
        ]
        
        summary_table = Table(summary_data, colWidths=[3*inch, 2*inch])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8B5CF6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ]))
        elements.append(summary_table)
        elements.append(Spacer(1, 0.3*inch))
        
        # Category breakdown
        if report_data['category_breakdown']:
            elements.append(Paragraph("Expense by Category", heading_style))
            category_data = [['Category', 'Total', 'Count', 'Percentage']]
            for category, data in report_data['category_breakdown'].items():
                category_data.append([
                    category,
                    f"${data['total']:,.2f}",
                    str(data['count']),
                    f"{data['percentage']:.1f}%"
                ])
            
            category_table = Table(category_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1.5*inch])
            category_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3B82F6')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(category_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Income source breakdown
        if report_data['source_breakdown']:
            elements.append(Paragraph("Income by Source", heading_style))
            source_data = [['Source', 'Total', 'Count', 'Percentage']]
            for source, data in report_data['source_breakdown'].items():
                source_data.append([
                    source,
                    f"${data['total']:,.2f}",
                    str(data['count']),
                    f"{data['percentage']:.1f}%"
                ])
            
            source_table = Table(source_data, colWidths=[2*inch, 1.5*inch, 1*inch, 1.5*inch])
            source_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#22C55E')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(source_table)
            elements.append(Spacer(1, 0.3*inch))
        
        # Top expenses
        if report_data['top_expenses']:
            elements.append(Paragraph("Top Expenses", heading_style))
            expense_data = [['Title', 'Amount', 'Category', 'Date']]
            for expense in report_data['top_expenses'][:10]:
                expense_data.append([
                    expense['title'][:30],
                    f"${expense['amount']:,.2f}",
                    expense['category'],
                    expense['date']
                ])
            
            expense_table = Table(expense_data, colWidths=[2.5*inch, 1.5*inch, 1.5*inch, 1*inch])
            expense_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#EF4444')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'RIGHT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 11),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightcoral),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ]))
            elements.append(expense_table)
        
        # Build PDF
        doc.build(elements)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def export_to_excel(report_data: Dict[str, Any]) -> bytes:
        """
        Convert report data to Excel format with multiple sheets
        
        Returns Excel bytes with sheets:
        - Summary
        - Category Breakdown
        - Income Sources
        - Top Expenses
        - Monthly Trends
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from io import BytesIO
        
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Summary sheet
        ws_summary = wb.create_sheet("Summary")
        ws_summary.append(["Financial Report Summary"])
        ws_summary.append([])
        ws_summary.append(["Period", f"{report_data['period']['start_date']} to {report_data['period']['end_date']}"])
        ws_summary.append(["Days", report_data['period']['days']])
        ws_summary.append(["Generated", report_data['generated_at']])
        ws_summary.append([])
        
        # Summary data
        summary = report_data['summary']
        ws_summary.append(["Metric", "Value"])
        ws_summary.append(["Total Income", summary['total_income']])
        ws_summary.append(["Total Expenses", summary['total_expenses']])
        ws_summary.append(["Balance", summary['balance']])
        ws_summary.append(["Income Transactions", summary['income_count']])
        ws_summary.append(["Expense Transactions", summary['expense_count']])
        ws_summary.append(["Avg Daily Expense", summary['avg_daily_expense']])
        ws_summary.append(["Avg Daily Income", summary['avg_daily_income']])
        
        # Style summary sheet
        ws_summary['A1'].font = Font(size=16, bold=True, color="8B5CF6")
        for row in ws_summary['A7:B14']:
            for cell in row:
                if cell.column == 1:
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(start_color="E0E7FF", end_color="E0E7FF", fill_type="solid")
        
        # Category breakdown sheet
        ws_category = wb.create_sheet("Category Breakdown")
        ws_category.append(["Category", "Total", "Count", "Percentage"])
        for category, data in report_data['category_breakdown'].items():
            ws_category.append([
                category,
                data['total'],
                data['count'],
                data['percentage']
            ])
        
        # Style category sheet
        for cell in ws_category[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Income sources sheet
        ws_income = wb.create_sheet("Income Sources")
        ws_income.append(["Source", "Total", "Count", "Percentage"])
        for source, data in report_data['source_breakdown'].items():
            ws_income.append([
                source,
                data['total'],
                data['count'],
                data['percentage']
            ])
        
        # Style income sheet
        for cell in ws_income[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="22C55E", end_color="22C55E", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Top expenses sheet
        ws_expenses = wb.create_sheet("Top Expenses")
        ws_expenses.append(["Title", "Amount", "Category", "Date", "Description"])
        for expense in report_data['top_expenses']:
            ws_expenses.append([
                expense['title'],
                expense['amount'],
                expense['category'],
                expense['date'],
                expense.get('description', '')
            ])
        
        # Style expenses sheet
        for cell in ws_expenses[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="EF4444", end_color="EF4444", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Monthly trends sheet
        ws_trends = wb.create_sheet("Monthly Trends")
        ws_trends.append(["Month", "Income", "Expenses", "Balance"])
        for trend in report_data['monthly_trends']:
            ws_trends.append([
                trend['month'],
                trend['income'],
                trend['expenses'],
                trend['balance']
            ])
        
        # Style trends sheet
        for cell in ws_trends[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="8B5CF6", end_color="8B5CF6", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")
        
        # Auto-adjust column widths for all sheets
        for ws in wb.worksheets:
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to bytes
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()
    
    @staticmethod
    def get_quick_report(
        user_id: int,
        period: str,
        db: Session
    ) -> Dict[str, Any]:
        """
        Generate quick report for predefined periods
        
        Periods:
        - this_month: Current month
        - last_month: Previous month
        - this_year: Current year
        - last_year: Previous year
        - last_30_days: Last 30 days
        - last_90_days: Last 90 days
        """
        today = date.today()
        
        if period == "this_month":
            start_date = date(today.year, today.month, 1)
            end_date = today
        elif period == "last_month":
            first_day_this_month = date(today.year, today.month, 1)
            end_date = first_day_this_month - timedelta(days=1)
            start_date = date(end_date.year, end_date.month, 1)
        elif period == "this_year":
            start_date = date(today.year, 1, 1)
            end_date = today
        elif period == "last_year":
            start_date = date(today.year - 1, 1, 1)
            end_date = date(today.year - 1, 12, 31)
        elif period == "last_30_days":
            start_date = today - timedelta(days=30)
            end_date = today
        elif period == "last_90_days":
            start_date = today - timedelta(days=90)
            end_date = today
        else:
            raise ValueError(f"Invalid period: {period}")
        
        return ReportService.generate_report(user_id, start_date, end_date, db)
